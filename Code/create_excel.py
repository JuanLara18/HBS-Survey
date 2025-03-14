import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import openpyxl

# Definir rutas
input_dir = "../Output/Results"
figures_dir = "../Output/Figures"
output_excel = "../Output/Training_Analysis_Results.xlsx"

# Asegurar que el directorio de salida existe
os.makedirs(os.path.dirname(output_excel), exist_ok=True)

# Función para dar formato a las hojas de Excel
def format_worksheet(ws, title):
    # Aplicar formato al título
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Formato para las cabeceras
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    
    return 3  # Devuelve la fila donde empezar a insertar datos

# Función para ajustar el ancho de las columnas
# Modify the adjust_column_width function to handle merged cells
def adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = None
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Skip merged cells
                try:
                    if column is None:
                        column = cell.column_letter  # Get column letter from first non-merged cell
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        if column:  # Only adjust if we found a valid column
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

# Función para aplicar estilo a una tabla
def apply_table_style(ws, start_row, end_row, start_col, end_col):
    # Formato de encabezado
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Aplicar estilo a la cabecera
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Aplicar bordes a toda la tabla
    border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if row > start_row:  # No es cabecera
                if row % 2 == 0:  # Filas pares
                    cell.fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")

# Función para insertar una tabla de DataFrame en una hoja
def insert_df_table(ws, df, start_row, start_col=1, include_index=False):
    # Insertar cabeceras
    for i, col_name in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + i)
        cell.value = col_name
        
    # Insertar datos
    for i, row in enumerate(df.values):
        for j, value in enumerate(row):
            cell = ws.cell(row=start_row + i + 1, column=start_col + j)
            # Formatear valores numéricos correctamente
            if isinstance(value, (int, float)):
                if isinstance(value, int):
                    cell.value = value
                elif abs(value) < 0.001:
                    cell.value = 0
                else:
                    # Usar 2 decimales para la mayoría de los números
                    cell.value = value
                    cell.number_format = '0.00'
            else:
                cell.value = value
    
    # Calcular el número de filas y columnas
    end_row = start_row + len(df)
    end_col = start_col + len(df.columns) - 1
    
    # Aplicar estilo a la tabla
    apply_table_style(ws, start_row, end_row, start_col, end_col)
    
    return end_row + 2  # Devuelve la siguiente fila disponible

# Función para insertar imagen en una hoja
def insert_image(ws, image_path, cell_ref):
    if os.path.exists(image_path):
        img = Image(image_path)
        # Redimensionar imagen si es necesario
        scale_factor = 0.75  # Ajustar según necesidad
        img.width = int(img.width * scale_factor)
        img.height = int(img.height * scale_factor)
        ws.add_image(img, cell_ref)
        return True
    else:
        print(f"Advertencia: No se encontró la imagen {image_path}")
        return False

# Crear un nuevo archivo Excel
wb = Workbook()

# PARTE 1: DESCRIBIENDO LA ENCUESTA
#============================================================================

# Hoja 1: Estadísticas resumidas de empresas
sheet1 = wb.active
sheet1.title = "Firm Summary Stats"
row_idx = format_worksheet(sheet1, "Summary Statistics of Firms")

# Intentar cargar datos de empresas
try:
    firm_size = pd.read_csv(f"{input_dir}/firm_size_stats.csv")
    sheet1.cell(row=row_idx, column=1).value = "Firm Size Distribution"
    sheet1.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    row_idx = insert_df_table(sheet1, firm_size, row_idx)
except Exception as e:
    print(f"Error loading firm size data: {e}")
    row_idx += 2

try:
    position_stats = pd.read_csv(f"{input_dir}/position_stats.csv")
    sheet1.cell(row=row_idx, column=1).value = "Position Distribution"
    sheet1.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    row_idx = insert_df_table(sheet1, position_stats, row_idx)
except Exception as e:
    print(f"Error loading position data: {e}")
    row_idx += 2

try:
    industry_stats = pd.read_csv(f"{input_dir}/industry_stats.csv")
    sheet1.cell(row=row_idx, column=1).value = "Industry Distribution"
    sheet1.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    row_idx = insert_df_table(sheet1, industry_stats, row_idx)
except Exception as e:
    print(f"Error loading industry data: {e}")
    row_idx += 2

# Hoja 2: Estadísticas resumidas sobre prácticas de RR.HH.
sheet2 = wb.create_sheet("HR Practices")
row_idx = format_worksheet(sheet2, "Summary Statistics on HR Practices")

# Cargar datos de prácticas de RR.HH.
try:
    hr_stats = pd.read_csv(f"{input_dir}/hr_practices_stats.csv")
    sheet2.cell(row=row_idx, column=1).value = "HR Practices Metrics"
    sheet2.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    row_idx = insert_df_table(sheet2, hr_stats, row_idx)
except Exception as e:
    print(f"Error loading HR practices data: {e}")
    row_idx += 2

# Insertar imágenes de prácticas de RR.HH.
row_idx += 1
sheet2.cell(row=row_idx, column=1).value = "HR Practices Visualizations"
sheet2.cell(row=row_idx, column=1).font = Font(bold=True)
row_idx += 1

# Intentar insertar imágenes relacionadas con HR
hr_images = [
    {"path": f"{figures_dir}/SummaryStats/HR/hr_workforce.png", "title": "Workforce Planning"},
    {"path": f"{figures_dir}/SummaryStats/HR/hr_skdescription.png", "title": "Skill Taxonomy"},
    {"path": f"{figures_dir}/SummaryStats/HR/hr_cat_skinventory.png", "title": "Skill Inventory"}
]

for i, img_info in enumerate(hr_images):
    img_path = img_info["path"]
    if os.path.exists(img_path):
        cell_ref = f"{get_column_letter(1)}{row_idx}"
        sheet2.cell(row=row_idx, column=1).value = img_info["title"]
        sheet2.cell(row=row_idx, column=1).font = Font(bold=True)
        row_idx += 1
        insert_image(sheet2, img_path, f"{get_column_letter(1)}{row_idx}")
        row_idx += 20  # Espacio para la imagen
    else:
        print(f"No se encontró la imagen {img_path}")

# Hoja 3: Estadísticas resumidas sobre brechas de habilidades
sheet3 = wb.create_sheet("Skills Gaps")
row_idx = format_worksheet(sheet3, "Summary Statistics on Skills Gaps and Coping Mechanisms")

# Cargar datos de brechas de habilidades
try:
    skills_gaps = pd.read_csv(f"{input_dir}/skills_gaps_stats.csv")
    sheet3.cell(row=row_idx, column=1).value = "Skills Gap Metrics"
    sheet3.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    row_idx = insert_df_table(sheet3, skills_gaps, row_idx)
except Exception as e:
    print(f"Error loading skills gaps data: {e}")
    row_idx += 2

# Insertar imágenes de brechas de habilidades
row_idx += 1
sheet3.cell(row=row_idx, column=1).value = "Skills Gap Visualizations"
sheet3.cell(row=row_idx, column=1).font = Font(bold=True)
row_idx += 1

# Intentar insertar imágenes de brechas de habilidades
skill_images = [
    {"path": f"{figures_dir}/skills_gap_severity.png", "title": "Severity of Skills Gap"},
    {"path": f"{figures_dir}/skills_most_needed.png", "title": "Skills Most Needed"},
    {"path": f"{figures_dir}/skills_by_org_type.png", "title": "Skills by Organization Type"},
    {"path": f"{figures_dir}/coping_mechanisms.png", "title": "Coping Mechanisms for Skills Gaps"},
    {"path": f"{figures_dir}/talent_sourcing_by_cluster.png", "title": "Talent Sourcing by Cluster"}
]

for i, img_info in enumerate(skill_images):
    img_path = img_info["path"]
    if os.path.exists(img_path):
        cell_ref = f"{get_column_letter(1)}{row_idx}"
        sheet3.cell(row=row_idx, column=1).value = img_info["title"]
        sheet3.cell(row=row_idx, column=1).font = Font(bold=True)
        row_idx += 1
        insert_image(sheet3, img_path, f"{get_column_letter(1)}{row_idx}")
        row_idx += 20  # Espacio para la imagen
    else:
        print(f"No se encontró la imagen {img_path}")

# PARTE 2: ANÁLISIS EN PROFUNDIDAD DEL ENTRENAMIENTO
#============================================================================

# Hoja 4: Fracción de entrenamiento genérico, upskilling y reskilling
sheet4 = wb.create_sheet("Training Types")
row_idx = format_worksheet(sheet4, "Fraction of Generic Training, Upskilling and Reskilling")

# Cargar datos de tipos de entrenamiento
try:
    training_types = pd.read_csv(f"{input_dir}/training_type_distribution.csv")
    sheet4.cell(row=row_idx, column=1).value = "Training Type Distribution"
    sheet4.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    row_idx = insert_df_table(sheet4, training_types, row_idx)
except Exception as e:
    print(f"Error loading training type data: {e}")
    row_idx += 2

# Insertar imágenes de distribución de entrenamiento
row_idx += 1
sheet4.cell(row=row_idx, column=1).value = "Training Type Visualizations"
sheet4.cell(row=row_idx, column=1).font = Font(bold=True)
row_idx += 1

# Intentar insertar imágenes
training_images = [
    {"path": f"{figures_dir}/training_distribution.png", "title": "Training Program Distribution (Bar)"},
    {"path": f"{figures_dir}/training_distribution_pie.png", "title": "Training Program Distribution (Pie)"}
]

for i, img_info in enumerate(training_images):
    img_path = img_info["path"]
    if os.path.exists(img_path):
        cell_ref = f"{get_column_letter(1)}{row_idx}"
        sheet4.cell(row=row_idx, column=1).value = img_info["title"]
        sheet4.cell(row=row_idx, column=1).font = Font(bold=True)
        row_idx += 1
        insert_image(sheet4, img_path, f"{get_column_letter(1)}{row_idx}")
        row_idx += 20  # Espacio para la imagen
    else:
        print(f"No se encontró la imagen {img_path}")

# Hoja 5: Estadísticas resumidas de características del programa
sheet5 = wb.create_sheet("Program Characteristics")
row_idx = format_worksheet(sheet5, "Summary Statistics of Program Characteristics")

# Cargar datos de características del programa
try:
    program_all = pd.read_csv(f"{input_dir}/program_characteristics_all.csv")
    sheet5.cell(row=row_idx, column=1).value = "Program Characteristics (All Programs)"
    sheet5.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    row_idx = insert_df_table(sheet5, program_all, row_idx)
except Exception as e:
    print(f"Error loading program characteristics (all) data: {e}")
    row_idx += 2

try:
    program_by_type = pd.read_csv(f"{input_dir}/program_characteristics_by_type.csv")
    sheet5.cell(row=row_idx, column=1).value = "Program Characteristics by Type (Upskilling vs Reskilling)"
    sheet5.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    row_idx = insert_df_table(sheet5, program_by_type, row_idx)
except Exception as e:
    print(f"Error loading program characteristics by type data: {e}")
    row_idx += 2

# Insertar imágenes de características del programa
row_idx += 1
sheet5.cell(row=row_idx, column=1).value = "Program Characteristics Visualizations"
sheet5.cell(row=row_idx, column=1).font = Font(bold=True)
row_idx += 1

# Intentar insertar imágenes
characteristic_images = [
    {"path": f"{figures_dir}/program_characteristics_comparison.png", "title": "Program Characteristics Comparison"},
    {"path": f"{figures_dir}/p_duration_by_program.png", "title": "Duration by Program Type"},
    {"path": f"{figures_dir}/p_hourstrained_by_program.png", "title": "Hours Trained by Program Type"},
    {"path": f"{figures_dir}/p_cost_by_program.png", "title": "Cost by Program Type"}
]

for i, img_info in enumerate(characteristic_images):
    img_path = img_info["path"]
    if os.path.exists(img_path):
        cell_ref = f"{get_column_letter(1)}{row_idx}"
        sheet5.cell(row=row_idx, column=1).value = img_info["title"]
        sheet5.cell(row=row_idx, column=1).font = Font(bold=True)
        row_idx += 1
        insert_image(sheet5, img_path, f"{get_column_letter(1)}{row_idx}")
        row_idx += 20  # Espacio para la imagen
    else:
        print(f"No se encontró la imagen {img_path}")

# Hoja 6: Clusters de programas de entrenamiento
sheet6 = wb.create_sheet("Training Clusters")
row_idx = format_worksheet(sheet6, "Clusters of Training Programs (2 Cluster Analysis)")

# Cargar datos de clusters
try:
    cluster_stats = pd.read_csv(f"{input_dir}/cluster_summary_stats.csv")
    sheet6.cell(row=row_idx, column=1).value = "Training Program Cluster Characteristics"
    sheet6.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    row_idx = insert_df_table(sheet6, cluster_stats, row_idx)
except Exception as e:
    print(f"Error loading cluster summary stats: {e}")
    row_idx += 2

# Insertar imágenes de clusters
row_idx += 1
sheet6.cell(row=row_idx, column=1).value = "Training Program Cluster Visualizations"
sheet6.cell(row=row_idx, column=1).font = Font(bold=True)
row_idx += 1

# Intentar insertar imágenes
cluster_images = [
    {"path": f"{figures_dir}/participation_by_cluster.png", "title": "Participation by Cluster"},
    {"path": f"{figures_dir}/structure_by_cluster.png", "title": "Structure by Cluster"},
    {"path": f"{figures_dir}/funding_by_cluster.png", "title": "Funding by Cluster"}
]

for i, img_info in enumerate(cluster_images):
    img_path = img_info["path"]
    if os.path.exists(img_path):
        cell_ref = f"{get_column_letter(1)}{row_idx}"
        sheet6.cell(row=row_idx, column=1).value = img_info["title"]
        sheet6.cell(row=row_idx, column=1).font = Font(bold=True)
        row_idx += 1
        insert_image(sheet6, img_path, f"{get_column_letter(1)}{row_idx}")
        row_idx += 20  # Espacio para la imagen
    else:
        print(f"No se encontró la imagen {img_path}")

# Hoja 7: Tabulación de cluster vs dummy de reskilling
sheet7 = wb.create_sheet("Cluster vs Reskilling")
row_idx = format_worksheet(sheet7, "Tabulation of Cluster vs Reskilling Dummy")

# Cargar datos de contingencia
try:
    contingency = pd.read_csv(f"{input_dir}/cluster_program_contingency.csv")
    sheet7.cell(row=row_idx, column=1).value = "Cluster vs Program Type Contingency Table"
    sheet7.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    row_idx = insert_df_table(sheet7, contingency, row_idx)
except Exception as e:
    print(f"Error loading cluster program contingency data: {e}")
    row_idx += 2

try:
    reskilling_by_cluster = pd.read_csv(f"{input_dir}/reskilling_by_cluster.csv")
    sheet7.cell(row=row_idx, column=1).value = "Characteristics of Reskilling Programs by Cluster"
    sheet7.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    row_idx = insert_df_table(sheet7, reskilling_by_cluster, row_idx)
except Exception as e:
    print(f"Error loading reskilling by cluster data: {e}")
    row_idx += 2

# Insertar imágenes de la relación cluster vs reskilling
row_idx += 1
sheet7.cell(row=row_idx, column=1).value = "Cluster vs Reskilling Visualizations"
sheet7.cell(row=row_idx, column=1).font = Font(bold=True)
row_idx += 1

# Intentar insertar imágenes
crsk_images = [
    {"path": f"{figures_dir}/cluster_by_program.png", "title": "Cluster Composition by Program Type"},
    {"path": f"{figures_dir}/reskilling_features_by_cluster.png", "title": "Reskilling Features by Cluster"}
]

for i, img_info in enumerate(crsk_images):
    img_path = img_info["path"]
    if os.path.exists(img_path):
        cell_ref = f"{get_column_letter(1)}{row_idx}"
        sheet7.cell(row=row_idx, column=1).value = img_info["title"]
        sheet7.cell(row=row_idx, column=1).font = Font(bold=True)
        row_idx += 1
        insert_image(sheet7, img_path, f"{get_column_letter(1)}{row_idx}")
        row_idx += 20  # Espacio para la imagen
    else:
        print(f"No se encontró la imagen {img_path}")

# PARTE 3: CORRELATOS EMPRESARIALES DE RESKILLING
#============================================================================

# Hoja 8: Regresión con dummy de reskilling como variable dependiente
sheet8 = wb.create_sheet("Reskilling Regressions")
row_idx = format_worksheet(sheet8, "Regression with Reskilling Dummy as Dependent Variable")

# Cargar datos de regresión
try:
    reskilling_reg = pd.read_csv(f"{input_dir}/reskilling_regressions.csv")
    sheet8.cell(row=row_idx, column=1).value = "Reskilling Regression Results"
    sheet8.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    row_idx = insert_df_table(sheet8, reskilling_reg, row_idx)
except Exception as e:
    print(f"Error loading reskilling regression data: {e}")
    row_idx += 2

# Descripción de las variables
row_idx += 1
sheet8.cell(row=row_idx, column=1).value = "Variable Descriptions"
sheet8.cell(row=row_idx, column=1).font = Font(bold=True)
row_idx += 1

variable_desc = pd.DataFrame({
    "Variable": ["f_medium", "f_large", "f_mne", "f_public", "f_union50", "hr_indexD_total", "sk_gap_sev", "clusplot2"],
    "Description": [
        "Medium-sized firms (reference: small firms)",
        "Large-sized firms (reference: small firms)",
        "Multinational enterprise dummy",
        "Publicly listed company dummy",
        "High union presence (>50%)",
        "HR practices index (total score)",
        "Severe skills gap dummy",
        "Skill needs cluster typology"
    ]
})
row_idx = insert_df_table(sheet8, variable_desc, row_idx)

# Hoja 9: Regresión con cluster como variable dependiente
sheet9 = wb.create_sheet("Cluster Regressions")
row_idx = format_worksheet(sheet9, "Regression with Cluster as Dependent Variable")

# Cargar datos de regresión
try:
    cluster_reg = pd.read_csv(f"{input_dir}/cluster_regressions.csv")
    sheet9.cell(row=row_idx, column=1).value = "Cluster Regression Results"
    sheet9.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    row_idx = insert_df_table(sheet9, cluster_reg, row_idx)
except Exception as e:
    print(f"Error loading cluster regression data: {e}")
    row_idx += 2

# Hoja 10: Regresión con reskilling complejo como variable dependiente
sheet10 = wb.create_sheet("Complex Reskilling")
row_idx = format_worksheet(sheet10, "Regression with Complex Reskilling as Dependent Variable")

# Cargar datos de regresión
try:
    complex_reg = pd.read_csv(f"{input_dir}/complex_reskilling_regressions.csv")
    sheet10.cell(row=row_idx, column=1).value = "Complex Reskilling Regression Results"
    sheet10.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    row_idx = insert_df_table(sheet10, complex_reg, row_idx)
    
    # Descripción de variable dependiente
    row_idx += 1
    sheet10.cell(row=row_idx, column=1).value = "Note: Complex Reskilling is defined as programs that are both Reskilling (program type) AND belong to Cluster 2 (more intensive programs)"
    sheet10.cell(row=row_idx, column=1).font = Font(italic=True)
    
except Exception as e:
    print(f"Error loading complex reskilling regression data: {e}")
    row_idx += 2

# Crear una hoja índice al principio
index_sheet = wb.create_sheet("Index", 0)
row_idx = format_worksheet(index_sheet, "Training Analysis Results - Table of Contents")

# Crear índice
index_data = pd.DataFrame({
    "Sheet": [
        "1. Firm Summary Stats", 
        "2. HR Practices",
        "3. Skills Gaps",
        "4. Training Types",
        "5. Program Characteristics",
        "6. Training Clusters",
        "7. Cluster vs Reskilling",
        "8. Reskilling Regressions",
        "9. Cluster Regressions",
        "10. Complex Reskilling"
    ],
    "Description": [
        "Summary statistics of firms (size, position, industry)",
        "Summary statistics on HR practices (individual indices and overall index)",
        "Summary statistics on skills gaps and coping mechanisms",
        "Fraction of generic training, upskilling and reskilling",
        "Summary statistics of program characteristics (total and by type)",
        "Clusters of training programs (2-cluster analysis)",
        "Tabulation of cluster vs reskilling dummy",
        "Regression with reskilling dummy as dependent variable",
        "Regression with cluster as dependent variable",
        "Regression with combined reskilling & complex program as dependent variable"
    ]
})

row_idx = insert_df_table(index_sheet, index_data, row_idx)

# Ajustar anchos de columna en todas las hojas
for sheet in wb.worksheets:
    adjust_column_width(sheet)

# Guardar el archivo Excel
wb.save(output_excel)
print(f"Excel file has been created: {output_excel}")